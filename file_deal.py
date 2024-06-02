import os
import time
from openpyxl import Workbook
import linecache
from file_class import File
from info_class import InfoClass


def sava_data_xlsx(ps_info_list: list):
    '''
    结果写入文件
    :param ps_info_list:文件对象list
    :return:
    '''
    # 新建工作簿
    wb = Workbook()
    for file in ps_info_list:
        ps_file_path = file.file_path_name
        print("正在将以下文件中的分析结果写入结果文件：")
        print(ps_file_path)
        # 每个ps文件数据创建一个sheet工作表保存
        ps_name = '_'.join((ps_file_path.split('\\')[-1]).split("_")[:2])
        ws = wb.create_sheet(ps_name,-1)
        # 最上面先写入一行这个ps文件的路径，标识这些数据的来源文件
        c = ws.cell(row=1, column=1)
        c.value = ps_file_path
        title_info = InfoClass()
        ws.append(title_info.get_title_info())
        # 如果结果数据列表不为空就写入数据
        # 为空就写入无数据结果
        if file.info_list:
            for infoclass in file.info_list:
            # for infoclass in (file.info_list).values():
                ws.append(infoclass.get_all_info())
        else:
            ws.append(["未查找到相关数据"])

    # 获取当前时间戳
    timestamp = str(int(time.time()))
    # 结果文件保存路径
    result_dir = os.getcwd()+'\\'+'result'
    if not os.path.exists(result_dir):
        os.mkdir(result_dir)
    file_path_name = os.path.join(os.getcwd()+r'\result','result_'+timestamp+'.xlsx')
    # 如果结果文件存在，将原来相同名称的文件通过时间戳进行改名备份，然后再新建文件
    # 不存在就直接新建
    if os.path.exists(file_path_name):
        os.rename(file_path_name, file_path_name + '_bak_' + timestamp)  # 将原文件重命名备份
    # 保存最终文件
    wb.save(file_path_name)
    print("*****文件分析完成*****")
    print("结果文件路径：" + file_path_name)


def find_ps_files_list(file_path: str) -> list:
    '''
    查找目标路径下包含的所有的ps文件
    :param file_path:
    :return:
    '''
    walk_generator = os.walk(file_path)
    file_root_path = {}
    ps_paths = []
    # 查看出目标路径下的所有文件和对应的上级路径，得到 {“文件夹路径”：“文件名”}
    for root_path, dirs, all_files in walk_generator:
        if len(all_files) > 0:
            file_root_path[root_path] = all_files

    # 根据得到的{“文件夹路径”：“文件名”}，查找出每个文件夹路径下的ps文件，并拼接成完整文件路径返回 [“ps_file_path”]
    for key,val in file_root_path.items():
        ps_files = [file for file in val if "_ps.dat" in file and "_ps.dat." not in file ]
        if ps_files:
            file = File()
            file.file_path_name = os.path.join(key,ps_files[0])
            ps_paths.append(file)
            # ps_paths.append(os.path.join(key,ps_files[0]))

    return ps_paths


def get_dsp_pattern_info(dsp_file_path):
    '''
    获取dsp文件中所有符合规则的行
    :param dsp_file_path:
    :return:
    '''
    contents = linecache.getlines(dsp_file_path)
    dsp_result_info = [line for line in contents if "DlshcedDesc be exist" in line or "dlBurstMeasResu" in line]
    linecache.clearcache()
    return dsp_result_info


def get_ps_pattern_info(ps_file_path,ps_pattern_list):
    '''
    获取ps文件中所有符合目标匹配关键字的所有行
    :param ps_file_path:
    :param ps_pattern_list:
    :return:
    '''
    contents = linecache.getlines(ps_file_path)
    # d1 = any(pattern if pattern in line else False for pattern in ps_pattern_list)
    ps_result_info = [line for line in contents for pattern in ps_pattern_list if pattern in line]
    linecache.clearcache()
    return ps_result_info


def uptade_dsp_rssi_sinr(psfile_result_list: list):
    '''
    查找dsp，更新rssi和sinr
    :param psfile_result_list:文件对象list
    :return:
    '''
    for file in psfile_result_list:
        ps_file_path = file.file_path_name
        print("开始校验下方文件中的数据：")
        print(ps_file_path)
        dsp_file_path = ps_file_path.replace("_ps.dat", "_dsp.dat")
        dsp_contents = get_dsp_pattern_info(dsp_file_path)
        for infoclass in file.info_list:
            # 只查找更新下行数据中rssi==-201的
            if infoclass.ul_dl_type == 'dl' and infoclass.rssi == -201:
                status = 0
                for i, line in enumerate(dsp_contents):
                    if status == 0 and "DlshcedDesc be exist" in line and str(infoclass.fn) in line:
                        status = 1
                    elif status == 1 and "dlBurstMeasResu" in line:
                        result = line.split("\t")
                        rssi = result[5].strip()
                        sinr = result[-1].strip()
                        infoclass.rssi = int(rssi)
                        infoclass.sinr = int(sinr)
                        break
                    else:
                        continue


def get_file_contents(file_path: str) -> list:
    '''
    读取文件所有内容到内存中，提高文件内容处理速度
    :param file_path:
    :return:
    '''
    contents = linecache.getlines(file_path)
    return contents


def close_file_cache():
    '''
    清楚读取到内存中的文件内容
    :return:
    '''
    linecache.clearcache()

