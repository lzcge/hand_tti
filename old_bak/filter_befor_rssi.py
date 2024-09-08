import os
import time
from openpyxl import Workbook
import linecache


def sava_data_xlsx(dsp_info_dict: dict):
    '''
    结果写入文件
    :param ps_info_list:文件对象list
    :return:
    '''
    # 新建工作簿
    wb = Workbook()
    for file_path,info_list in dsp_info_dict.items():
        print("正在将以下文件中的分析结果写入结果文件：")
        print(file_path)
        # 每个ps文件数据创建一个sheet工作表保存
        ps_name = '_'.join((file_path.split('\\')[-1]).split("_")[:2])
        ws = wb.create_sheet(ps_name,-1)
        # 最上面先写入一行这个ps文件的路径，标识这些数据的来源文件
        c = ws.cell(row=1, column=1)
        c.value = file_path
        ws.append(["fn","tsn","phy_agc1","phy_agc2","phy_filter_befor_rssi"])
        # 如果结果数据列表不为空就写入数据
        # 为空就写入无数据结果
        if info_list:
            for info in info_list:
                ws.append(info)
        else:
            ws.append(["未查找到相关数据"])

    # 获取当前时间戳
    timestamp = str(int(time.time()))
    # 结果文件保存路径
    result_dir = os.getcwd()+'\\'+'result'
    if not os.path.exists(result_dir):
        os.mkdir(result_dir)
    file_path_name = os.path.join(os.getcwd()+r'\result','result_'+timestamp+'.xlsx')
    # 如果结果文件存在，将原来相同名称的文件通过时间戳进行改名备份，然后再新建文件
    # 不存在就直接新建
    if os.path.exists(file_path_name):
        os.rename(file_path_name, file_path_name + '_bak_' + timestamp)  # 将原文件重命名备份
    # 保存最终文件
    wb.save(file_path_name)
    print("*****文件分析完成*****")
    print("结果文件路径：" + file_path_name)


def find_dsp_files_list(file_path: str) -> list:
    '''
    查找目标路径下包含的所有的ps文件
    :param file_path:
    :return:
    '''
    walk_generator = os.walk(file_path)
    file_root_path = {}
    dsp_paths = []
    # 查看出目标路径下的所有文件和对应的上级路径，得到 {“文件夹路径”：“文件名”}
    for root_path, dirs, all_files in walk_generator:
        if len(all_files) > 0:
            file_root_path[root_path] = all_files

    # 根据得到的{“文件夹路径”：“文件名”}，查找出每个文件夹路径下的ps文件，并拼接成完整文件路径返回 [“ps_file_path”]
    for key,val in file_root_path.items():
        ps_files = [file for file in val if "_dsp.dat" in file]
        if ps_files:

            file_path_name = os.path.join(key,ps_files[0])
            dsp_paths.append(file_path_name)
            # ps_paths.append(os.path.join(key,ps_files[0]))

    return dsp_paths


def get_dsp_pattern_info(dsp_file_path):
    '''
    获取dsp文件中所有符合规则的行
    :param dsp_file_path:
    :return:
    '''
    contents = linecache.getlines(dsp_file_path)
    dsp_result_info = [line for line in contents if "dl burst sche" in line or "dlBurstAgcAve" in line or "dlBurstMeasResu" in line]
    linecache.clearcache()
    return dsp_result_info


def get_dsp_filter_befor_rssi(dsp_result_info: list):
    '''
    查找dsp，更新rssi和sinr
    :param psfile_result_list:文件对象list
    :return:
    '''
    result = []
    fn, tsn,phy_agc1,phy_agc2,phy_filter_befor_rssi = -1,-1,-1,-1,-1
    # 当前帧标志位，标志有dl burst sche的打印的帧才提取保存后面的dlBurstAgcAve信息，防止重复打印dlBurstAgcAve时重复保存
    current_fn_start_status = 0
    for i,line in enumerate(dsp_result_info):
        if "dl burst sche" in line:
            fns = line.split("\t")
            fn = int(fns[5].strip())
            tsn = int(fns[6].strip())
            current_fn_start_status = 1
        if "dlBurstAgcAve" in line and current_fn_start_status == 1:
            phy_filter_befor_rssis = line.split("\t")
            phy_agc1 = int(phy_filter_befor_rssis[5].strip())
            phy_agc2 = int(phy_filter_befor_rssis[6].strip())
            phy_filter_befor_rssi = int(phy_filter_befor_rssis[7].strip())
            result.append([fn, tsn,phy_agc1,phy_agc2,phy_filter_befor_rssi])
            current_fn_start_status = 0


    '''
    其中key表示按第几个关键字排序，lambda x:后()中的元素表示了关键字的优先级
    第一位表示最高的优先级，第二位次之。。。
    其中下x[0]表示按第一个关键词的升序排列，-x[1]表示按第二个关键词的降序排列。
    '''
    # result.sort(key=lambda x: (x[0], -x[1]))
    # 按帧号升序排序
    result.sort(key=lambda x: x[0])
    return result



if __name__ == '__main__':
    time1 = time.time()
    dsp_file_paths = find_dsp_files_list(r"C:\Users\lenovo\Desktop\每日分析\外场分析\00给应用院提交的文件\99 临时其它内容\下行功率抖动图汇总\20240826新方案\20240906-5682圈次-重庆-早上\20240906-TC5682圈次-重庆-凌晨-移动通信-采数测试\UE40-MO-USRP采数\20240906_044029486")
    dsp_result_dict = {}
    for file in dsp_file_paths:
        all_lines = get_dsp_pattern_info(file)
        result_list = get_dsp_filter_befor_rssi(all_lines)
        dsp_result_dict[file] = result_list
    sava_data_xlsx(dsp_result_dict)
