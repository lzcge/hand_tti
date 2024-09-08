import os
import time
from openpyxl import Workbook
import linecache
from common.file_class import File
from common.info_class import InfoClass
from common.search_keyword_enum import SearchKeyword
from common.pattern_extract_class import PatternExtract


def sava_data_xlsx(ps_info_list: list):
    '''
    结果写入文件
    :param ps_info_list:文件对象list
    :return:
    '''
    # 新建工作簿
    wb = Workbook()
    for file in ps_info_list:
        ps_file_path = file.file_path_name
        print("正在将以下文件中的分析结果写入结果文件：")
        print(ps_file_path)
        # 每个ps文件数据创建一个sheet工作表保存
        ps_name = '_'.join((ps_file_path.split('\\')[-1]).split("_")[:2])
        ws = wb.create_sheet(ps_name, -1)
        # 最上面先写入一行这个ps文件的路径，标识这些数据的来源文件
        c = ws.cell(row=1, column=1)
        c.value = ps_file_path
        ws.append(InfoClass.get_infoclass_all_attribute())
        # 如果结果数据列表不为空就写入数据
        # 为空就写入无数据结果
        if file.info_list:
            for infoclass in file.info_list:
                # for infoclass in (file.info_list).values():
                ws.append(infoclass.get_all_info())

        else:
            ws.append(["未查找到相关数据"])

    # 获取当前时间戳
    timestamp = str(int(time.time()))
    # 结果文件保存路径
    result_dir = os.getcwd() + '\\' + 'result'
    if not os.path.exists(result_dir):
        os.mkdir(result_dir)
    file_path_name = os.path.join(os.getcwd() + r'\result', 'result_' + timestamp + '.xlsx')
    # 如果结果文件存在，将原来相同名称的文件通过时间戳进行改名备份，然后再新建文件
    # 不存在就直接新建
    if os.path.exists(file_path_name):
        os.rename(file_path_name, file_path_name + '_bak_' + timestamp)  # 将原文件重命名备份
    # 保存最终文件
    wb.save(file_path_name)
    print("*****文件分析完成*****")
    print("结果文件路径：" + file_path_name)


def find_ps_files_list(file_path: str) -> list:
    '''
    查找目标路径下包含的所有的ps文件
    :param file_path:
    :return:
    '''
    walk_generator = os.walk(file_path)
    file_root_path = {}
    ps_paths = []
    # 查看出目标路径下的所有文件和对应的上级路径，得到 {“文件夹路径”：“文件名”}
    for root_path, dirs, all_files in walk_generator:
        if len(all_files) > 0:
            file_root_path[root_path] = all_files

    # 根据得到的{“文件夹路径”：“文件名”}，查找出每个文件夹路径下的ps文件，并拼接成完整文件路径返回 [“ps_file_path”]
    for key, val in file_root_path.items():
        ps_files = [file for file in val if "_ps.dat" in file and "_ps.dat." not in file]
        if ps_files:
            file = File()
            file.file_path_name = os.path.join(key, ps_files[0])
            ps_paths.append(file)
            # ps_paths.append(os.path.join(key,ps_files[0]))

    return ps_paths


def get_file_keywords_lines(file_path, search_keywords):
    """
    查找一个文件中包含关键词的所有行
    :param file_path: 文件路径
    :param search_keywords: 查找关键词
    :return:
    """
    file_all_contents = linecache.getlines(file_path)
    file_keywords_contents = [line for line in file_all_contents for keyword in search_keywords if keyword in line]
    linecache.clearcache()
    return file_keywords_contents


def get_dsp_all_filter_befor_rssi(dsp_keywords_contents: list):
    """
    获取物理层日志中所有滤波前功率
    :param dsp_keywords_contents: 物理层滤波前功率根据关键字提取后的所有行信息
    :return:
    """
    result = []
    fn, tsn, phy_agc1, phy_agc2, phy_filter_befor_rssi = -1, -1, -1, -1, -1
    # 当前帧标志位，标志有dl burst sche的打印的帧才提取保存后面的dlBurstAgcAve信息，防止重复打印dlBurstAgcAve时重复保存
    current_fn_start_status = 0
    for i, line in enumerate(dsp_keywords_contents):
        if SearchKeyword.PHY_DL_BURST_SCHE.value in line:
            fns = line.split("\t")
            fn = int(fns[5].strip())
            tsn = int(fns[6].strip())
            current_fn_start_status = 1
        if SearchKeyword.PHY_DLBURSTAGCAVE.value in line and current_fn_start_status == 1:
            phy_filter_befor_rssis = line.split("\t")
            phy_agc1 = int(phy_filter_befor_rssis[5].strip())
            phy_agc2 = int(phy_filter_befor_rssis[6].strip())
            phy_filter_befor_rssi = int(phy_filter_befor_rssis[7].strip())
            result.append([fn, tsn, phy_agc1, phy_agc2, phy_filter_befor_rssi])
            current_fn_start_status = 0

    '''
    其中key表示按第几个关键字排序，lambda x:后()中的元素表示了关键字的优先级
    第一位表示最高的优先级，第二位次之。。。
    其中下x[0]表示按第一个关键词的升序排列，-x[1]表示按第二个关键词的降序排列。
    '''
    # result.sort(key=lambda x: (x[0], -x[1]))
    # 按帧号升序排序
    result.sort(key=lambda x: x[0])
    return result

# @staticmethod
# def uptade_dsp_rssi_sinr(psfile_result_list: list):
#     '''
#     查找dsp，更新rssi和sinr
#     :param psfile_result_list:文件对象list
#     :return:
#     '''
#     for file in psfile_result_list:
#         ps_file_path = file.file_path_name
#         print("开始校验下方文件中的数据：")
#         print(ps_file_path)
#         dsp_file_path = ps_file_path.replace("_ps.dat", "_dsp.dat")
#         dsp_contents = get_dsp_pattern_info(dsp_file_path, ["DlshcedDesc be exist", "dlBurstMeasResu"])
#         for infoclass in file.info_list:
#             # 只查找更新下行数据中rssi==-201的
#             if infoclass.ul_dl_type == 'dl' and infoclass.rssi == -201 and infoclass.fn:
#                 status = 0
#                 for i, line in enumerate(dsp_contents):
#                     if status == 0 and "DlshcedDesc be exist" in line and str(infoclass.fn) in line:
#                         status = 1
#                     elif status == 1 and "dlBurstMeasResu" in line:
#                         result = line.split("\t")
#                         rssi = result[5].strip()
#                         sinr = result[-1].strip()
#                         infoclass.rssi = int(rssi)
#                         infoclass.sinr = int(sinr)
#                         break
#                     else:
#                         continue


def find_start_statellite(contents: str) -> int:
    '''
    查找文件中第一个出现的最开始的卫星id
    :param pattern_rule:
    :param contents:
    :return:
    '''
    # 查找一开始的卫星ID初始值
    for i, statellite_line in enumerate(contents):
        if SearchKeyword.SATELLITE.value in statellite_line:
            satellite = PatternExtract.satellite_pattern_search(statellite_line)
            return satellite
        else:
            continue


def get_file_contents(file_path: str) -> list:
    '''
    读取文件所有内容到内存中，提高文件内容处理速度
    :param file_path:
    :return:
    '''
    contents = linecache.getlines(file_path)
    return contents


def close_file_cache():
    '''
    清楚读取到内存中的文件内容
    :return:
    '''
    linecache.clearcache()


def merge_dsp_filter_befor_rssi(dsp_file_path: str,file: File):
    """
    把dsp中的滤波前功率补充到ps结果里
    :param dsp_file_path:
    :param file:
    :return:
    """
    # 查找当前文件中的物理层滤波前功率，按帧、时隙和ps中的数据进行合并
    # dsp_file_path = ps_file_path.replace("_ps.dat", "_dsp.dat")
    dsp_keywords_contents = get_file_keywords_lines(dsp_file_path, SearchKeyword.get_dsp_search_keywords())
    dsp_all_filter_befor_rssi = get_dsp_all_filter_befor_rssi(dsp_keywords_contents)
    dsp_all_filter_befor_rssi_fns = []
    dsp_all_filter_befor_rssi_tsns = []
    for filter_befor_rssi_item in dsp_all_filter_befor_rssi:
        dsp_all_filter_befor_rssi_fns.append(str(filter_befor_rssi_item[0]))
        dsp_all_filter_befor_rssi_tsns.append(str(filter_befor_rssi_item[1]))

    current_file_ps_result_infolist = file.info_list
    for index in range(len(current_file_ps_result_infolist)):
        current_infoclass = current_file_ps_result_infolist[index]
        if current_infoclass.ul_dl_type != 'dl':
            continue
        elif str(current_infoclass.fn) in dsp_all_filter_befor_rssi_fns:
            current_dsp_fn_index = dsp_all_filter_befor_rssi_fns.index(str(current_infoclass.fn))
            # current_dsp_fn = dsp_all_filter_befor_rssi_fns[current_dsp_fn_index]
            # current_dsp_tsn = dsp_all_filter_befor_rssi_tsns[current_dsp_fn_index]

            current_dsp_filter_befor_agc1 = dsp_all_filter_befor_rssi[current_dsp_fn_index][2]
            current_dsp_filter_befor_agc2 = dsp_all_filter_befor_rssi[current_dsp_fn_index][3]
            current_dsp_filter_befor_rssi = dsp_all_filter_befor_rssi[current_dsp_fn_index][4]
            current_infoclass.phy_agc1 = current_dsp_filter_befor_agc1
            current_infoclass.phy_agc2 = current_dsp_filter_befor_agc2
            current_infoclass.phy_filter_befor_rssi = current_dsp_filter_befor_rssi
        else:
            pass




